"""
save_workplan_status.py  —  Local Script สำหรับรันบนเครื่อง Windows
────────────────────────────────────────────────────────────────────
ดึงข้อมูลสถานะแผนงานจาก LINE HR Bot แล้วบันทึก Excel ลง HR folder

วิธีใช้:
  python save_workplan_status.py
  python save_workplan_status.py --date 2026-06-29
  python save_workplan_status.py --url https://xxxx.up.railway.app

ตั้งเวลาอัตโนมัติด้วย Windows Task Scheduler:
  ดู setup_windows_task.bat ในโฟลเดอร์นี้
"""
import argparse
import json
import os
import sys
from datetime import date, datetime

import httpx
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─── Config ───────────────────────────────────────────────────────────────────
# แก้ค่าเหล่านี้ให้ตรงกับ setup ของคุณ

BOT_BASE_URL   = os.getenv("BOT_BASE_URL", "https://xxxx.up.railway.app")   # URL ของ Railway bot
SAVE_FOLDER    = os.getenv("WORKPLAN_SAVE_PATH", r"D:\Watercourse\Department\HR\WorkPlan")
REQUEST_TIMEOUT = 15  # วินาที


# ─── Helper ───────────────────────────────────────────────────────────────────

def get_workplan_status(target_date: str) -> dict:
    """ดึงข้อมูลจาก Bot API endpoint /workplan-status"""
    url = f"{BOT_BASE_URL}/workplan-status"
    try:
        resp = httpx.get(url, params={"date": target_date}, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        return resp.json()
    except httpx.HTTPError as e:
        print(f"[ERROR] ไม่สามารถเชื่อมต่อ Bot API: {e}")
        print(f"        ตรวจสอบว่า BOT_BASE_URL ถูกต้อง: {BOT_BASE_URL}")
        sys.exit(1)


def next_revision(folder: str, date_prefix: str) -> int:
    """หา revision ถัดไป ตามไฟล์ที่มีอยู่แล้วในวันนั้น"""
    if not os.path.exists(folder):
        return 1
    existing = [
        f for f in os.listdir(folder)
        if f.startswith(date_prefix) and "WorkPlanStatus" in f and f.endswith(".xlsx")
    ]
    return len(existing) + 1


def build_excel(data: dict, filepath: str):
    """สร้างไฟล์ Excel สวยงามด้วย openpyxl"""
    target_date = data.get("date", "")
    submitted   = set(data.get("submitted", []))
    not_sub     = data.get("not_submitted", [])
    all_staff   = data.get("all_staff", [])
    checked_at  = data.get("checked_at", datetime.now().strftime("%H:%M น."))
    total       = len(all_staff) or len(submitted) + len(not_sub)

    # ─── สร้าง rows ───
    rows = []
    for name in sorted(all_staff):
        rows.append({
            "ลำดับ":       len(rows) + 1,
            "ชื่อพนักงาน": name,
            "สถานะ":       "✅ ส่งแล้ว" if name in submitted else "❌ ยังไม่ส่ง",
        })
    for name in sorted(submitted - set(all_staff)):
        rows.append({
            "ลำดับ":       len(rows) + 1,
            "ชื่อพนักงาน": name,
            "สถานะ":       "✅ ส่งแล้ว",
        })
    if not rows:
        rows.append({"ลำดับ": "-", "ชื่อพนักงาน": "ยังไม่มีข้อมูลวันนี้", "สถานะ": "-"})

    df = pd.DataFrame(rows)

    # ─── สร้าง Excel ───
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # metadata sheet
        meta_rows = [
            ["วันที่",         target_date],
            ["เวลาตรวจ",       checked_at],
            ["พนักงานทั้งหมด", total],
            ["ส่งแผนงานแล้ว",  len(submitted)],
            ["ยังไม่ส่ง",      len(not_sub)],
            ["% ส่งแล้ว",      f"{len(submitted)/total*100:.0f}%" if total else "-"],
        ]
        pd.DataFrame(meta_rows, columns=["รายการ", "ค่า"]).to_excel(
            writer, index=False, sheet_name="สรุป"
        )
        df.to_excel(writer, index=False, sheet_name="รายชื่อ")

        wb = writer.book

        # ─── จัดสไตล์ Sheet "สรุป" ───
        ws_sum = writer.sheets["สรุป"]
        _style_summary_sheet(ws_sum, len(submitted), len(not_sub), total)

        # ─── จัดสไตล์ Sheet "รายชื่อ" ───
        ws_det = writer.sheets["รายชื่อ"]
        _style_detail_sheet(ws_det, submitted)

    print(f"[OK] บันทึกแล้ว: {filepath}")


def _style_summary_sheet(ws, sent: int, not_sent: int, total: int):
    from openpyxl.styles import PatternFill, Font, Alignment

    navy   = PatternFill("solid", fgColor="1E3C72")
    green  = PatternFill("solid", fgColor="C6EFCE")
    red    = PatternFill("solid", fgColor="FFC7CE")
    white  = Font(color="FFFFFF", bold=True, size=12)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # header row
    for cell in ws[1]:
        cell.fill = navy
        cell.font = white

    # แถวสถิติสำคัญ
    for row in ws.iter_rows(min_row=2):
        label = str(row[0].value or "")
        if "ส่งแผนงานแล้ว" in label:
            for c in row:
                c.fill = green
        elif "ยังไม่ส่ง" in label:
            for c in row:
                c.fill = red

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.row_dimensions[1].height = 22


def _style_detail_sheet(ws, submitted: set):
    from openpyxl.styles import PatternFill, Font, Alignment

    navy  = PatternFill("solid", fgColor="1E3C72")
    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")
    alt   = PatternFill("solid", fgColor="EFF3FB")
    white = Font(color="FFFFFF", bold=True, size=12)

    # header
    for cell in ws[1]:
        cell.fill = navy
        cell.font = white
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # data rows
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        name_cell = row[1] if len(row) > 1 else None
        status_cell = row[2] if len(row) > 2 else None
        if status_cell and "✅" in str(status_cell.value or ""):
            fill = green
        elif status_cell and "❌" in str(status_cell.value or ""):
            fill = red
        else:
            fill = alt if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="บันทึกสถานะแผนงาน LINE HR Bot")
    parser.add_argument("--date", default=date.today().isoformat(), help="YYYY-MM-DD")
    parser.add_argument("--url",  default=BOT_BASE_URL, help="Bot base URL")
    args = parser.parse_args()

    global BOT_BASE_URL
    BOT_BASE_URL = args.url.rstrip("/")

    print(f"[INFO] ดึงข้อมูลวันที่ {args.date} จาก {BOT_BASE_URL}")

    data = get_workplan_status(args.date)

    os.makedirs(SAVE_FOLDER, exist_ok=True)
    date_prefix = args.date.replace("-", "")
    rev         = next_revision(SAVE_FOLDER, date_prefix)
    filename    = f"{date_prefix}_WorkPlanStatus_Rev.{rev}.xlsx"
    filepath    = os.path.join(SAVE_FOLDER, filename)

    build_excel(data, filepath)

    # สรุปผล
    sent    = len(data.get("submitted", []))
    not_sub = len(data.get("not_submitted", []))
    print(f"\n📊 สรุป: ส่งแล้ว {sent} คน | ยังไม่ส่ง {not_sub} คน")
    if data.get("not_submitted"):
        print("❌ ยังไม่ส่ง:")
        for name in data["not_submitted"]:
            print(f"   • {name}")


if __name__ == "__main__":
    main()

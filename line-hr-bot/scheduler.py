"""
scheduler.py
APScheduler jobs อัตโนมัติ 2 งาน:
  1. 09:30 น. — ตรวจแผนงาน → บันทึก Excel เท่านั้น (ไม่แจ้ง LINE)
  2. 09:45 น. — ส่ง Daily Report รูปภาพเข้า LINE กลุ่ม
"""
import asyncio
import logging
import os
from datetime import datetime, timezone, timedelta

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

from config import REPORT_TIME, WORKPLAN_DEADLINE, LINE_GROUP_ID, WORKPLAN_SAVE_PATH

logger = logging.getLogger(__name__)

# ใช้ UTC+7 offset สำหรับแสดงเวลาไทย (ไม่ใช้ timezone library)
_THAI_TZ = timezone(timedelta(hours=7))

# Scheduler รัน UTC — ไม่ต้องการ pytz หรือ zoneinfo
_scheduler = BackgroundScheduler()


def _now_thai() -> datetime:
    """เวลาปัจจุบันตามเวลาไทย"""
    return datetime.now(_THAI_TZ)


def _today_thai() -> str:
    """คืน YYYY-MM-DD ตามเวลาไทย"""
    return datetime.now(_THAI_TZ).strftime("%Y-%m-%d")


def _thai_to_utc(hhmm: str) -> tuple[int, int]:
    """แปลงเวลาไทย HH:MM → UTC hour, minute (ลบ 7 ชั่วโมง)"""
    h, m = map(int, hhmm.split(":"))
    return (h - 7) % 24, m


# ──────────────────────────────────────────────────────────────────────────────
# JOB 1 : ตรวจ 09:30 น. — ใครยังไม่ส่งแผนงาน
# ──────────────────────────────────────────────────────────────────────────────

def _run_workplan_check():
    today = _today_thai()
    logger.info(f"[SCHEDULER] ตรวจแผนงาน {today} {WORKPLAN_DEADLINE} น.")
    try:
        asyncio.run(_check_and_notify(today))
    except Exception as e:
        logger.error(f"[SCHEDULER workplan_check] {e}")


async def _check_and_notify(today: str):
    """09:30 น. — บันทึก Excel เท่านั้น ไม่ส่งแจ้งใน LINE"""
    from database import get_messages_by_date, get_attendance_by_date

    messages   = get_messages_by_date(today)
    attendance = get_attendance_by_date(today)

    submitted = {
        m["sender_name"]
        for m in messages
        if m["category"] == "work_plan"
    }
    all_staff = [
        a["employee_name"]
        for a in attendance
        if a["status"] in ("present", "late")
    ]
    not_submitted = [name for name in all_staff if name not in submitted]

    # บันทึก Excel ลง WORKPLAN_SAVE_PATH — ไม่ส่งอะไรเข้า LINE
    await _save_workplan_excel(today, submitted, not_submitted, all_staff)
    logger.info(f"[WORKPLAN CHECK] ส่งแล้ว={len(submitted)}, ยังไม่ส่ง={len(not_submitted)} (บันทึก Excel เท่านั้น)")


async def _save_workplan_excel(today: str, submitted: set, not_submitted: list, all_staff: list):
    """บันทึกสถานะแผนงานเป็น Excel ลง WORKPLAN_SAVE_PATH"""
    import pandas as pd
    from datetime import datetime

    try:
        os.makedirs(WORKPLAN_SAVE_PATH, exist_ok=True)

        # หมายเลข Revision (ถ้าไฟล์วันนี้มีแล้ว ให้เพิ่ม Rev)
        date_prefix = today.replace("-", "")
        existing = [
            f for f in os.listdir(WORKPLAN_SAVE_PATH)
            if f.startswith(date_prefix) and f.endswith(".xlsx")
        ]
        rev = len(existing) + 1
        filename = f"{date_prefix}_WorkPlanStatus_Rev.{rev}.xlsx"
        filepath = os.path.join(WORKPLAN_SAVE_PATH, filename)

        # สร้าง DataFrame
        rows = []
        for name in sorted(all_staff):
            rows.append({
                "ลำดับ": len(rows) + 1,
                "ชื่อพนักงาน": name,
                "สถานะ": "✅ ส่งแล้ว" if name in submitted else "❌ ยังไม่ส่ง",
                "เวลาตรวจ": _now_thai().strftime("%H:%M น."),
                "วันที่": today,
            })
        # พนักงานที่ส่งแต่ไม่อยู่ใน HumanSoft
        extra = sorted(submitted - set(all_staff))
        for name in extra:
            rows.append({
                "ลำดับ": len(rows) + 1,
                "ชื่อพนักงาน": name,
                "สถานะ": "✅ ส่งแล้ว (ไม่มีใน HumanSoft)",
                "เวลาตรวจ": _now_thai().strftime("%H:%M น."),
                "วันที่": today,
            })

        if not rows:
            rows.append({
                "ลำดับ": "-",
                "ชื่อพนักงาน": "ยังไม่มีข้อมูลวันนี้",
                "สถานะ": "-",
                "เวลาตรวจ": _now_thai().strftime("%H:%M น."),
                "วันที่": today,
            })

        df = pd.DataFrame(rows)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="สถานะแผนงาน")
            ws = writer.sheets["สถานะแผนงาน"]

            # ปรับความกว้าง column
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)

            # สีแถว header
            from openpyxl.styles import PatternFill, Font, Alignment
            header_fill = PatternFill("solid", fgColor="1E3C72")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

        logger.info(f"[WORKPLAN EXCEL] บันทึก: {filepath}")

    except Exception as e:
        logger.error(f"[WORKPLAN EXCEL] save error: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# JOB 2 : Daily Report 09:45 น.
# ──────────────────────────────────────────────────────────────────────────────

def _run_daily_report():
    from generators.report_generator import generate_and_send_report
    today = _today_thai()
    logger.info(f"[SCHEDULER] เริ่มสร้าง report วันที่ {today}")
    try:
        asyncio.run(generate_and_send_report(LINE_GROUP_ID, today))
    except Exception as e:
        logger.error(f"[SCHEDULER daily_report] {e}")


# ──────────────────────────────────────────────────────────────────────────────

def start_scheduler():
    # แปลงเวลาไทย → UTC (Railway รัน UTC server)
    # Thailand = UTC+7, ดังนั้น 09:30 Thai = 02:30 UTC, 09:45 Thai = 02:45 UTC
    wp_utc_h, wp_m = _thai_to_utc(WORKPLAN_DEADLINE)
    rp_utc_h, rp_m = _thai_to_utc(REPORT_TIME)

    # Job 1 — ตรวจแผนงาน (UTC)
    _scheduler.add_job(
        _run_workplan_check,
        CronTrigger(hour=wp_utc_h, minute=wp_m),  # UTC — ไม่ต้องระบุ timezone
        id="workplan_check",
        replace_existing=True,
    )

    # Job 2 — Daily Report (UTC)
    _scheduler.add_job(
        _run_daily_report,
        CronTrigger(hour=rp_utc_h, minute=rp_m),  # UTC — ไม่ต้องระบุ timezone
        id="daily_report",
        replace_existing=True,
    )

    _scheduler.start()
    logger.info(
        f"[SCHEDULER] ✅ ตั้งเวลา (UTC → Thai):\n"
        f"  • {wp_utc_h:02d}:{wp_m:02d} UTC = {WORKPLAN_DEADLINE} น. — บันทึก Excel แผนงาน\n"
        f"  • {rp_utc_h:02d}:{rp_m:02d} UTC = {REPORT_TIME} น. — ส่ง Daily Report เข้า LINE กลุ่ม"
    )
